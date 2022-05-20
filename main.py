import openpyxl
import requests
from bs4 import BeautifulSoup
import selenium
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from time import sleep

from bokjiro_crawler import BokjiroCrawler

def writeData(sheet, data):
  sheet.cell(row=1, column=1).value = 'index'
  index = 2
  for key in data[0]:
    sheet.cell(row=1, column=index).value = key
    index += 1

  row = 2
  for item in data:
    sheet.cell(row=row, column = 1).value = row - 1
    column = 2
    for val in item.values():
      temp = val
      if type(val) is list:
        temp = ','.join(val)
      sheet.cell(row=row, column = column).value = temp
      column += 1
    row +=1



crawler = BokjiroCrawler()
crawler.driverInit()


# wb = openpyxl.Workbook()
# central_data = crawler.getCentralData()
# sheet1 = wb.active
# sheet1.title = '중앙부처'
# writeData(sheet1, central_data)
# wb.save('crawler_central.xlsx')

wb = openpyxl.Workbook()
local_data = crawler.getLocalData()
sheet1 = wb.active
sheet1.title = '지자체'
writeData(sheet1, local_data)
wb.save('crawler_local.xlsx')

# wb = openpyxl.Workbook()
# private_data = crawler.getPrivateData()
# sheet1 = wb.active
# sheet1.title = '민간'
# writeData(sheet1, private_data)
# wb.save('crawler_private.xlsx')



# wb.get_

# url = "https://www.bokjiro.go.kr/ssis-teu/twataa/wlfareInfo/moveTWAT52005M.do"
# # url = "https://www.bokjiro.go.kr/ssis-teu/twataa/wlfareInfo/moveTWAT52015M.do?wlfareInfoId=BOK00000762&wlfareInfoReldBztpCd=03"

# driver = webdriver.Chrome(executable_path='./chromedriver')
# driver.get(url=url)

# driver.implicitly_wait(time_to_wait=5)


# def getCurrentIndex():
#   element = driver.find_elements(by=By.CSS_SELECTOR, value='.cl-pageindexer-index.cl-selected')
#   return element.text

# def test():
#   current_index  = getCurrentIndex()

#   # self.moveToNextPage()
#   sleep(1)

# test()
#   # changed_index = self.getCurrentIndex()

#   # while changed_index != current_index:
#   #   current_index = changed_index
#   #   self.moveToNextPage()
#   #   sleep(1)
#   #   changed_index = self.getCurrentIndex()


# def getCurrentLastIndex(self):
#   pagination_elements = self.driver.find_elements(by=By.CLASS_NAME, value='cl-pageindexer-index')
#   return pagination_elements.pop().text

# def moveToNextPaginationIndex(self):
#   button = self.driver.find_element(by=By.CLASS_NAME, value='cl-pageindexer-next')
#   button.click()

# def moveToNextPage(self):
#   buttons = self.driver.find_element(by=By.CLASS_NAME, value='cl-pageindexer-index')
#   button.click()
# print(text[1])

# temp = text[1].find_elements(by=By.CSS_SELECTOR, value='.cl-last-row > .cl-output > div')
# print(temp)
# # filter_buttons = driver.find_elements(by=By.CLASS_NAME,value='cl-text-wrapper')

# # 장애인 태그 클릭
# for button in filter_buttons:
#   # print(button.accessible_name)
#   if button.accessible_name == '장애인':
#     button.click()
#     break

# # 복지 관련 내용 검색
# is_find = False

# for button in filter_buttons:
#   if button.aria_role == 'button':
#     inner_elements = button.find_elements(by=By.CLASS_NAME, value='cl-text')
#     for element in inner_elements:
#       if element.text == '검색':
#           element.click()
#           is_find = True
#           break
#   if is_find == True:
#     break


# sleep(3) # 3초 대기


# # # 카드 타이틀 가져오기

# test = driver.find_elements(by=By.CLASS_NAME, value='tabfolder-item')
# print(test)

# # cards = driver.find_elements(by=By.CLASS_NAME,value='card-tit')

# # for card in cards:
# #   if card.accessible_name != '':
# #     card_elements = card.find_elements(by=By.)
#   # print(card)


# # test = driver.find_elements(by=By.CLASS_NAME, value='badge')

# cards = []

# cards_elements = driver.find_elements(by=By.CSS_SELECTOR, value='.cl-layout-content > .card')
# for card_element in cards_elements:
#   card = {}
#   # print(card_element.text)
#   badges = []
#   badges_elements = card_element.find_elements(by=By.CSS_SELECTOR, value='.badge')
#   for badge_element in badges_elements:
#     # print('----------------------------')
#     # print(badge_element.text)
#     badges.append(badge_element.text)
#   # print('----------------------------')
#   card['badges'] = badges
  
#   title_element = card_element.find_element(by=By.CSS_SELECTOR, value='.card-tit')
#   card['title'] = title_element.text

#   sub_title_element = card_element.find_element(by=By.CSS_SELECTOR, value='.card-subtit')
#   card['sub_title'] = sub_title_element.text

#   sub_content_elements = card_element.find_elements(by=By.CSS_SELECTOR, value='.blt-tit ~ .cl-control')
#   card['contact'] = sub_content_elements[0].text
#   card['period'] = sub_content_elements[1].text
#   card['type'] = sub_content_elements[2].text
#   card['department'] = sub_content_elements[3].text

#   button_element = card_element.find_element(by=By.CSS_SELECTOR, value='.btn-secondary')
#   button_element.click()

#   sleep(3) # 3초 대기

#   #내부 페이지에서 tag별 정리
#   circle_tag_elements = driver.find_elements(by=By.CLASS_NAME,value='card-circle')
#   card['department_sub'] = circle_tag_elements[0].text
#   card['base_year'] = circle_tag_elements[3].text

#   support_target_element = driver.find_elements(by=By.CLASS_NAME,value='bokjiBlit01')
#   card['support_target'] = support_target_element[0].text
#   card['selection_criteria'] = support_target_element[1].text

#   tab_button_elements = driver.find_elements(by=By.CLASS_NAME, value='cl-tabfolder-item')

#   # 서비스 내용
#   tab_button_elements[1].click()
#   service_contents_elements = driver.find_elements(by=By.CLASS_NAME,value='bokjiBlit01')
#   card['service_contents'] = service_contents_elements[0].text

#   # 신청방법
#   tab_button_elements[2].click()
#   apply_method_elements = driver.find_elements(by=By.CLASS_NAME, value='process-v-bar')
#   card['apply_method'] = apply_method_elements[0].text

#   # 추가정보
#   tab_button_elements[3].click()
#   additional_info_elements = driver.find_elements(by=By.CLASS_NAME, value='process-v-bar')
#   additional_info = ''
#   for additional_info_element in additional_info_elements:
#     additional_info += additional_info_element.text + '\n'
#   card['additional_info'] = additional_info


#   driver.execute_script("window.history.go(-1)")


#   cards.append(card)

# print(cards)


# # search_box = driver.find_element(by=By.XPATH,value='//*[@id="uuid-b0d22cdd-7ab6-e78e-d187-5f9ea74c5d94"]/div[1]/input')
# # search_box = driver.find_element(by=By.CLASS_NAME,value='cl-text')
# # print(search_box)
